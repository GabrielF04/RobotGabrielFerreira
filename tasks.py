
import sys
import openpyxl # type: ignore
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from robocorp import workitems
import json
import re
from selenium.common.exceptions import StaleElementReferenceException
import os


class Scrappy:
    
    def __init__(self):
        self.driver = webdriver.Chrome()
        self.wait = WebDriverWait(self.driver, 50)
        self.search_phrase = None 
    
    def element_visible(self, p_Xpath):
        try:
            element = self.wait.until(EC.visibility_of_element_located((By.XPATH, p_Xpath)))
            return element
        except TimeoutException:
            print('element not work')
            return None
        
    def list_elements_visible(self, p_Xpath):
        try:
            elements = self.wait.until(EC.presence_of_all_elements_located((By.XPATH, p_Xpath)))
            return elements
        except TimeoutException:
            print('element not work')
            return None
          
    def initiating(self):
        ### call work item
        
        script_directory = os.path.dirname(__file__)
        
        file_path = os.path.join(script_directory, 'output/work-items-in/workitems.json')

        with open(file_path, 'r') as f:
            work_items = json.load(f)
        
        for work_item in work_items:
            payload = work_item["payload"]
            search_phrase = payload["search_phrase"]
            category = payload["category"]
            months = payload["months"]

        ### call the url ###
        self.driver.maximize_window()
        self.driver.get('https://www.latimes.com/')

        try:
            page = self.wait.until(EC.presence_of_element_located((By.XPATH, '//body')))
            print('Page load without error')
        except TimeoutException:
            print('Page not load')
            self.driver.quit()

        ### Interactions with the news in home page
        btn_field = '//button[@data-element="search-button"]'
        search_field = '//input[@class="w-full text-2xl leading-none border-0 text-secondary-color-7 md:text-4xl-1"]'
        btn_category = f'//li[@data-element="quick-links-item"]//a[@href="https://www.latimes.com/{category}"]'

        element_indentificated = self.element_visible(btn_category)
        if element_indentificated:
            element_indentificated.click()
            time.sleep(2)
        else: 
            print("button category not found")
            self.driver.quit() 

        element_indentificated = self.element_visible(btn_field)
        if element_indentificated:
            element_indentificated.click()
            time.sleep(2)
        else: 
            print("button search not found")
            self.driver.quit()

        element_indentificated = self.element_visible(search_field)    
        if element_indentificated:
            if search_phrase:  # Verifica se search_phrase foi atribuído
                element_indentificated.send_keys(search_phrase)
                time.sleep(2)
                element_indentificated.send_keys(Keys.RETURN)
                time.sleep(5)
            else: 
                print("search_phrase is not assigned")
                self.driver.quit()
        else: 
            print("search field not found")
            self.driver.quit()

        ### Interactions with the news in second page

        btn_sort_by = '//select[@class="select-input"]'

        element_indentificated = self.element_visible(btn_sort_by)
        if months == 2:
            if element_indentificated:
                element_indentificated.click()
                time.sleep(2)
                element_indentificated.send_keys(Keys.DOWN)
                time.sleep(2)
                element_indentificated.send_keys(Keys.RETURN)
        elif months == 1:
            if element_indentificated:
                element_indentificated.click()
                time.sleep(2)
                element_indentificated.send_keys(Keys.DOWN)
                time.sleep(2)
                element_indentificated.send_keys(Keys.DOWN)
                time.sleep(2)
                element_indentificated.send_keys(Keys.RETURN)
        else:
            print("Invalid value for 'months' parameter")
            self.driver.quit()

        ## I create the list and store the news data 
        list_title_news = []
        list_description_news = []
        list_image_news = []
        list_date_news = []
        
        regex_money = r'\b(?:R\$\s*\d+(?:[.,]\d{1,2})?|\d+(?:[.,]\d{1,2})?\s*(?:reais|dólares?))\b'

        title_news = self.list_elements_visible('//div[@class="promo-title-container"]')

        for title in title_news:
            try:
                list_title_news.append(title.text)
                title_money = title.text
                money_title = bool(re.search(regex_money, title_money))
                time.sleep(1)
            except StaleElementReferenceException:
                title_news = self.list_elements_visible('//div[@class="promo-title-container"]')

        description_news = self.list_elements_visible('//div[@class="promo-wrapper"]//p[@class="promo-description"]')
        list_description_news = []
        for description in description_news:
            list_description_news.append(description.text)
            description_money = description.text
            money_description = bool(re.search(regex_money, description_money))
        time.sleep(1)

        image_news = self.list_elements_visible('//div[@class="promo-media"]//a[@class="link promo-placeholder"]//img')
        list_image_news = []
        for image in image_news:
            list_image_news.append(image.get_attribute('src'))
        time.sleep(1)

        date_news = self.list_elements_visible('//div[@class="promo-wrapper"]//p[@class="promo-timestamp"]')
        list_date_news = []
        for date in date_news:
            list_date_news.append(date.text)
        time.sleep(10)
        print(list_image_news)

        self.create_workbook(list_title_news, list_description_news, list_image_news, list_date_news, money_title, money_description)

    def create_workbook(self, list_title_news, list_description_news, list_image_news, list_date_news, money_title, money_description):
        index = 2
        workbook = openpyxl.Workbook()
        news = workbook['Sheet']
        news.title = 'News'
        news['A1'] = 'Title'
        news['B1'] = 'Description'
        news['C1'] = 'Image link'
        news['D1'] = 'Date'
        news['E1'] = 'Title have money'
        news['F1'] = 'Description have money'

        for title, description, image, date in zip(list_title_news, list_description_news, list_image_news, list_date_news):
            news.cell(column=1, row=index, value=title)
            news.cell(column=2, row=index, value=description)
            news.cell(column=3, row=index, value=image)
            news.cell(column=4, row=index, value=date)
            news.cell(column=5, row=index, value=money_title)  
            news.cell(column=6, row=index, value=money_description)  
            index += 1
        workbook.save('Robot_news_excel.xlsx')
        print('Workbook has been created')
        sys.exit()
start = Scrappy()
start.initiating()
